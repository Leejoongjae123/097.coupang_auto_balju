import json

import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
# from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import pprint


def load_excel(fname,startNumber,endNumber):
    # fname = 'result.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(startNumber, endNumber+1):
        sellerNo = ws.cell(row=i, column=3).value
        if sellerNo == None:
            print('데이타 더 이상 없음')
            break
        receiverName = ws.cell(row=i, column=4).value
        receiverPhone = ws.cell(row=i, column=5).value
        basicAddress = ws.cell(row=i, column=6).value
        detailAddress = ws.cell(row=i, column=7).value
        productName=ws.cell(row=i, column=8).value
        productQuantity = int(ws.cell(row=i, column=9).value)
        url = ws.cell(row=i, column=11).value
        data = {'sellerNo':sellerNo,'receiverName':receiverName,'receiverPhone':receiverPhone,'basicAddress':basicAddress,'detailAddress':detailAddress,'productName':productName,'productQuantity':productQuantity,'url':url,'rowNum':i}
        data_list.append(data)
    print(data_list)
    return data_list
def chrome_browser(url):
    chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
    driver_path = f'./{chrome_ver}/chromedriver.exe'

    if os.path.exists(driver_path):
        print(f"chromedriver is installed: {driver_path}")
    else:
        print(f"install the chrome driver(ver: {chrome_ver})")
        chromedriver_autoinstaller.install(True)

    try:
        shutil.rmtree(r"c:\chrometemp")  # 쿠키 / 캐쉬파일 삭제
    except FileNotFoundError:
        pass

    subprocess.Popen(
        r'C:\Program Files\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\chrometemp"')  # 디버거 크롬 구동

    print("옵션설정")
    option = Options()
    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'
    option.add_argument('user-agent='+user_agent)
    option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    option.add_argument("--disable - blink - features = AutomationControlled")
    chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
    try:
        browser = webdriver.Chrome(driver_path, options=option)
        browser.maximize_window()
        url_start = url
        browser.get(url_start)
    except:
        chromedriver_autoinstaller.install(True)
        browser = webdriver.Chrome(options=option)
    return browser


import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db

def get_key(first_flag):
    if first_flag==True:
        secret={
          "type": "service_account",
          "project_id": "hanbaik",
          "private_key_id": "14ac8e67260883ee2c095ab58e32b27c41df23dd",
          "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDbX5A3/CzD78yf\n68NUAGB8fGyDAwY8wk2ngS2aYHNCNsnk2MVJQtCZTEo9PaRGG+xrlXKpJBMFQdmZ\nIctMpff9E+q0PDDASSgIjhbmcRT4vfMe2gRTeMISdezirXkbzGFKAKvjhQdtF+y2\n3y/KIpHDJeyesTKaAdOCAcFj2G/V3JucXNhxrrwgFhHHqBGET0c+BDTh/9DkX2M7\nav4hSmkdGGma2qsFdc1LzfWEwrN6Izqv4VGAFxSbhTL6/eyAxOc9jqbho3jVn9Ix\nXSpLsI64LwImJWPBCigy0Hoo9PMxyBtocKygDaK9Ii79KCF1cvIUaDZY1nu6Wns9\nknN21vR/AgMBAAECggEAOEFLVHAAcsZ9rMzirBnkpEer8/TElrQlAb1omlv9co9m\nOp80CFNv9r/PkeDKzYe/mt8aJjGwBvsZ9+Dop2EwNN/0of+FaOnZsEfvq2x4OklL\nnS+/SECBVAaVlpNxqSVCFZ4SdifY/arS8xpMwQMYafjBsvgxx7iMKpyUoSwRkb+l\nQtMagdhTEvJl7QwLnCpLgGFWs/3GjG8M5RD8CdHjGqxTaigYOkjU6bD00csIHkbI\nGveP585Qnf1rSbjZ83blk2S6gRqb9+OW4J8Z5Okwm4Dmb3LXkHgGCoVNT6f1sFsZ\nhsm1L8h7gztJwVmpoNTSDr4xQKiXQRGwb4KRfoHuMQKBgQDv730N4itlFadYxZdw\nktskJhjKs4N8LRCO6dy7uoU1GNIuiy+xFXhcGNSSnSM6JarbRZaVU0H4qtd9KtZ0\nl0fIO6jZLTQKyDJJEZifVYEKFXW811yfYBtvObh00gt40/26dajqct2oulIQxzq6\n7NVpa9EEdEaCRq2kTa+c8SI60wKBgQDqD6MdUgV5S3vmn4KW4uxT60ZtucvWT/H7\nOJNzvjT3YdeMhKG1gbB/wOn+W5tNMj5zzOH7Zngu+fCATHNU/ndohAp0cas3itw8\nIpIlNRR45Vvdcjn19Ml0qVh87oS9O6B9njCXC56jKKOAc5kWjgA6VsHJFEYBh8SU\nYe7TdZk8JQKBgEE9gWdxBBOsW6CLua3mgKfHpB4ZybrOFh6GAHsbMHVLlnsJZaJl\nECEar1JeX+HDtD2DInrf9KRE7+sc5ss1B1OuxS6oV+pGnUW4/yL0AO5Y/3alqI29\neDg6HanGI1BrdCZrL87wBM2IPCBLy/BfzXeo1WC8rR9nUHfIl+O4vXH5AoGBANgr\nDvuyV+nZRBoP8XzHIXqzzTzjnpVVCmh5rPz1i1d6HqfhirPmjgq/MZzAICNgpvsu\nGvujfJXuMidb9BxoVAHMCRfYL0hBz/sd9pm0dy7crUZNC6jTpgc/q8DeTOu0GRpL\nMhceHSoVC0RD/vwss5stqxW5ypn5OR3NgNP9RUOdAoGBALglxCsn4/O3aDB4CotP\nbzdQWSLneFXEUz3iQsTIvdWMUnT7onwWVkECD1DmcpngFy+/b2wEHp4uOrKvb5Nx\nLbrzpppHTBOmtEnh076T5K3JtP+ZgdmxgkjX3csRoQBk1VhdK+l1Cx1j9krct9rB\nLpDPBYw2Uyt/OucarZjmMf56\n-----END PRIVATE KEY-----\n",
          "client_email": "firebase-adminsdk-nkdhd@hanbaik.iam.gserviceaccount.com",
          "client_id": "103946102534201211018",
          "auth_uri": "https://accounts.google.com/o/oauth2/auth",
          "token_uri": "https://oauth2.googleapis.com/token",
          "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
          "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-nkdhd%40hanbaik.iam.gserviceaccount.com"
        }

        cred = credentials.Certificate(secret)
        firebase_admin.initialize_app(cred,{
            'databaseURL': "https://hanbaik-default-rtdb.firebaseio.com/"
        })

    password=db.reference().get()['users']['password']
    id = db.reference().get()['users']['id']
    print("id:",id,'password:',password)
    # ref=db.reference().get()
    # result_password=ref['users']

    return id,password

class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성
    user_signal2 = pyqtSignal(str)  # 사용자 정의 시그널 2 생성
    user_signal3 = pyqtSignal(int)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,auth,fname,startNumber,endNumber,id,pw):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.auth=auth
        self.fname=fname
        self.startNumber=startNumber
        self.endNumber=endNumber
        self.id=id
        self.pw=pw
        self.buyMore=False
    def run(self):
        if self.auth==True:
            print("작업시작")
            firstFlag = True
            fName = self.fname
            startNumber = self.startNumber
            endNumber = self.endNumber
            productList = load_excel(fName, startNumber, endNumber)
            wb = openpyxl.load_workbook(fName)
            ws = wb.active

            url_login = 'https://login.coupang.com/login/login.pang?rtnUrl=https%3A%2F%2Fwww.coupang.com%2Fnp%2Fpost%2Flogin%3Fr%3Dhttps%253A%252F%252Fwww.coupang.com%252F'
            browser = chrome_browser(url_login)

            # id='ljj3347@naver.com'
            # pw='dlwndwo2'
            id = self.id
            pw = self.pw

            browser.implicitly_wait(3)
            input_id = browser.find_element(By.ID, 'login-email-input')
            input_id.send_keys(id)
            time.sleep(0.5)

            input_pw = browser.find_element(By.ID, 'login-password-input')
            input_pw.send_keys(pw)
            time.sleep(0.5)

            btn_login = browser.find_element(By.CSS_SELECTOR,
                                             'body > div.member-wrapper.member-wrapper--flex > div.member-main > div.member-login._loginRoot.sms-login-target.style-v2 > form > div.login__content.login__content--trigger > button')
            browser.execute_script("arguments[0].click();", btn_login)  #

            while True:
                print("로그인시도중..")
                soup = BeautifulSoup(browser.page_source, 'lxml')
                isInputKeyword = len(soup.find_all('input', attrs={'id': 'headerSearchKeyword'}))
                if isInputKeyword >= 1:
                    print("로그인완료")
                    break
                time.sleep(1)

            for productElem in productList:
                url_purchase = productElem['url']
                browser.get(url_purchase)
                browser.implicitly_wait(3)

                text=f"{productElem['productName']} 구매중..."
                self.user_signal.emit(text)

                try:
                    btnUp = browser.find_element(By.CLASS_NAME, 'prod-quantity__plus')
                    actionBtn = productElem['productQuantity'] - 1
                    print("증가횟수:", actionBtn)
                    for i in range(0, actionBtn):
                        print(f"{i}번째 클릭")
                        btnUp = browser.find_element(By.CLASS_NAME, 'prod-quantity__plus')
                        btnUp.click()
                        time.sleep(3)
                except:
                    print("갯수 증가 안됨")

                # 구매가능갯수 파악하기
                soup = BeautifulSoup(browser.page_source, 'lxml')
                scripts = soup.find_all('script')
                result = ""
                for script in scripts:
                    if str(script).find("wishList") >= 0:
                        result = script
                        break

                rawscript = str(result)
                # print(rawscript)
                splitPositionFr = rawscript.find("=")
                splitPositionRr = rawscript.find(";")
                # print(splitPositionFr,splitPositionRr)
                rawscriptChanged = rawscript[splitPositionFr + 1:splitPositionRr].strip()
                # print(rawscriptChanged)
                jsonRawScript = json.loads(rawscriptChanged)
                # pprint.pprint(jsonRawScript)
                buyableQuantity = jsonRawScript['buyableQuantity']
                print("구매가능수량:", buyableQuantity)
                print("구매할수량:", productElem['productQuantity'])

                if buyableQuantity == None:
                    print('구매불가능')
                    ws.cell(row=productElem['rowNum'], column=10).value = "주문불가"
                    wb.save(fName)
                    text = f"{productElem['productName']} 구매 불가..."
                    self.user_signal.emit(text)
                    continue
                elif buyableQuantity == 0 or buyableQuantity < productElem['productQuantity']:
                    print('구매불가능')
                    text = f"{productElem['productName']} 구매 불가..."
                    self.user_signal.emit(text)
                    ws.cell(row=productElem['rowNum'], column=10).value = "주문불가"
                    wb.save(fName)
                    continue

                btnPurchase = browser.find_element(By.CLASS_NAME, 'prod-buy-btn__txt')
                btnPurchase.click()
                print("구매버튼클릭하기")
                browser.implicitly_wait(5)

                while True:
                    try:
                        btnAddress = browser.find_element(By.CSS_SELECTOR,
                                                          '#body > div.middle > div:nth-child(4) > h2 > button')
                        btnAddress.click()
                        time.sleep(1)
                        print("구매버튼누르기성공")
                        break
                    except:
                        print("구매버튼누르기실패")
                    time.sleep(1)

                while True:
                    print("창 뜨는것 대기, 창갯수:{}".format(len(browser.window_handles)))
                    if len(browser.window_handles) >= 2:
                        break
                    time.sleep(0.5)

                browser.switch_to.window(browser.window_handles[-1])
                print("마지막창으로 이동")

                btnModify = browser.find_element(By.CSS_SELECTOR,
                                                 'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div > div.address-card.address-card--picked > div.address-card__foot > form.address-card__form.address-card__form--edit._addressBookAddressCardEditForm > button')
                btnModify.click()
                browser.implicitly_wait(3)

                btnDelete = browser.find_element(By.CSS_SELECTOR,
                                                 'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div.content-body__corset > div > form > button')
                btnDelete.click()
                browser.implicitly_wait(3)

                interval = 1
                prev_height = browser.execute_script('return document.body.scrollHeight')
                while True:
                    browser.execute_script('window.scrollTo(0,document.body.scrollHeight)')
                    time.sleep(interval)
                    curr_height = browser.execute_script('return document.body.scrollHeight')
                    if curr_height == prev_height:
                        break
                        prev_height = curr_height

                btnAddAddress = browser.find_element(By.CSS_SELECTOR,
                                                     'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div > form > div > button')
                btnAddAddress.click()
                # browser.execute_script("arguments[0].click();", btnAddAddress)  #
                time.sleep(1)





                inputName = browser.find_element(By.ID, 'addressbookRecipient')
                inputName.send_keys(productElem['receiverName'])
                time.sleep(0.5)
                print("이름입력")

                inputPhone = browser.find_element(By.ID, 'addressbookCellphone')
                inputPhone.send_keys(productElem['receiverPhone'])
                time.sleep(0.5)
                print("전화번호입력")

                searchAddress = browser.find_element(By.CSS_SELECTOR,
                                                     'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div.content-body__corset > form > div.icon-text-field__frame-box._addressBookAddressErrorStatus > div > div.icon-text-field__button-container > a')
                searchAddress.click()
                browser.implicitly_wait(3)
                print("서칭하기2")

                soup = BeautifulSoup(browser.page_source, 'lxml')
                # print(soup.prettify())

                element = browser.find_element(By.CLASS_NAME, "identity__iframe")
                browser.switch_to.frame(element)  # 프레임 이동

                inputAddressBasic = browser.find_element(By.NAME, 'searchKey')
                inputAddressBasic.send_keys(productElem['basicAddress'])
                time.sleep(0.5)
                print("기본주소정보입력")

                btnSearch = browser.find_element(By.CSS_SELECTOR,
                                                 'body > section > div.zipcode__wrapper > div > div > header > div > form > div.zipcode__search-trigger > button')
                browser.execute_script("arguments[0].click();", btnSearch)  #
                browser.implicitly_wait(3)
                print("서치버튼누르기3")
                btnRowAddress = browser.find_element(By.CSS_SELECTOR,
                                                     'body > section > div.zipcode__wrapper > div > div > div > div.zipcode__slide-view._zipcodeResultSlideRoot > div.zipcode__slide-track._zipcodeResultSlide > div.zipcode__slide-item.zipcode__slide-item--address._zipcodeResultSlideItem > div._zipcodeResultListAddress > div:nth-child(1) > span.zipcode__result__item.zipcode__result__item--road._zipcodeResultSendTrigger')
                browser.execute_script("arguments[0].click();", btnRowAddress)  #
                browser.implicitly_wait(3)
                print("첫번째행선택")

                browser.switch_to.default_content()
                soup = BeautifulSoup(browser.page_source, 'lxml')
                # print(soup.prettify())
                inputDetail = browser.find_element(By.CSS_SELECTOR, '#addressbookAddressDetail')
                inputDetail.send_keys(productElem['detailAddress'])
                time.sleep(0.5)

                checkBase = browser.find_element(By.ID, '_addressBookSaveAsDefault')
                browser.execute_script("arguments[0].click();", checkBase)  #
                time.sleep(0.5)

                try:
                    print("와우인듯")
                    deliveryType=browser.find_element(By.CSS_SELECTOR,'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div.content-body__corset > form > div.icon-text-field__frame-box._addressBookDeliveryPreferencesErrorStatus > div > div.icon-text-field__button-container._addressBookDawnDeliveryPreferences > a')
                    deliveryType.click()
                    time.sleep(1)

                    element = browser.find_element(By.CLASS_NAME, "identity__iframe")
                    browser.switch_to.frame(element)  # 프레임 이동

                    elem = WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div > form > div.preference-required-dawn__radio-text > div.preference-required.__AA04_REQUEST_FOD_ADDITIONAL_EASY_ACCESS_LOBBY.main_picker__child_radio > label > span')))
                    deliveryFree=browser.find_element(By.CSS_SELECTOR,'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div > form > div.preference-required-dawn__radio-text > div.preference-required.__AA04_REQUEST_FOD_ADDITIONAL_EASY_ACCESS_LOBBY.main_picker__child_radio > label > span')
                    deliveryFree.click()
                    time.sleep(1)

                    agreeAndSave=browser.find_element(By.CSS_SELECTOR,'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div > form > button')
                    browser.execute_script("arguments[0].click();", agreeAndSave)  #
                    time.sleep(0.5)

                    browser.switch_to.default_content()
                except:
                    print("와우아닌듯")



                btnSave = browser.find_element(By.CSS_SELECTOR,
                                               'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div.content-body__corset > form > div.addressbook__button-fixer > button')
                browser.execute_script("arguments[0].click();", btnSave)  #
                time.sleep(0.5)




                browser.switch_to.window(browser.window_handles[0])

                if firstFlag == True:
                    paymentType = browser.find_element(By.CSS_SELECTOR, '#payType9')
                    browser.execute_script("arguments[0].click();", paymentType)  #
                    time.sleep(0.5)

                    btnChangeInfo = browser.find_element(By.CSS_SELECTOR,
                                                         '#body > div.middle > div:nth-child(8) > div > div > button')
                    btnChangeInfo.click()
                    time.sleep(0.5)

                    selectOption = browser.find_element(By.CSS_SELECTOR,
                                                        '#body > div.middle > div:nth-child(8) > div > div > div:nth-child(2) > div.cash-receipt__resiter-type__wrap > span:nth-child(1) > select')
                    selectOption.click()
                    time.sleep(0.5)

                    selectSaupja = browser.find_element(By.CSS_SELECTOR,
                                                        '#body > div.middle > div:nth-child(8) > div > div > div:nth-child(2) > div.cash-receipt__resiter-type__wrap > span:nth-child(1) > select > option:nth-child(2)')
                    selectSaupja.click()
                    time.sleep(0.5)

                    inputSaupja = browser.find_element(By.CSS_SELECTOR,
                                                       '#body > div.middle > div:nth-child(8) > div > div > div:nth-child(2) > div.cash-receipt__resiter-type__wrap > span:nth-child(2) > input')

                    inputSaupja.click()
                    time.sleep(0.5)
                    ActionChains(browser).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
                    time.sleep(0.5)
                    ActionChains(browser).send_keys('delete')

                    inputSaupja.send_keys(productElem['sellerNo'])
                firstFlag = False

                btnPay = browser.find_element(By.CSS_SELECTOR, '#paymentBtn')
                browser.execute_script("arguments[0].click();", btnPay)  #
                successFlag = False
                successCount = 10
                while True:
                    soup = BeautifulSoup(browser.page_source, 'lxml')
                    if successCount >= 2:
                        print("조회 한도 초과로 실패")
                        break
                    # print(soup.prettify())
                    try:
                        payStatus = soup.find('span', attrs={'class': 'i18n-wrapper'})
                        payStatusText = payStatus.get_text()
                        print('payStatusText:', payStatusText)
                        if payStatusText.find("완료") >= 0:
                            print("주문완료됨")
                            successFlag = True
                            break
                    except:
                        print(f"아직안뜸_{successCount}")
                    successCount += 1

                    time.sleep(1)

                if successFlag == True:
                    text = f"{productElem['productName']} 구매 완료..."
                    self.user_signal.emit(text)
                    print("주문완료로 엑셀에 저장")
                    ws.cell(row=productElem['rowNum'], column=10).value = "주문성공"
                else:
                    text = f"{productElem['productName']} 구매 불가..."
                    self.user_signal.emit(text)
                    print("주문실패로 엑셀에 저장")
                    ws.cell(row=productElem['rowNum'], column=10).value = "주문불가"
                wb.save(fName)

                text="또 구매할까요?"
                print(text)
                self.user_signal2.emit(text)
                self.buyMore=False
                self.escape=False

                while True:
                    print("추가구매여부확인중...")
                    if self.buyMore==True:
                        if self.escape==False:
                            text = "추가 구매 시작"
                            self.user_signal.emit(text)
                            break
                        else:
                            text = "추가 구매 중지"
                            print(text)
                            self.user_signal.emit(text)
                            break
                    time.sleep(1)
                if self.escape==True:
                    print('추가 구매 중지됨')
                    self.user_signal3.emit(int(productElem['rowNum']))
                    browser.close()
                    break
            text="작업 완료"
            self.user_signal.emit(text)
        else:
            text="로그인 문제 있음"
            print(text)
            self.user_signal.emit(text)

    def More(self):
        self.buyMore=True
        self.escape=False

    def Escape(self):
        self.buyMore = True
        self.escape=True
    def stop(self):
        pass

class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        QApplication.processEvents()
        self.lineEdit.setText("hanbaik0422")
        self.lineEdit_2.setEchoMode(QLineEdit.Password)
        self.lineEdit_2.setText("gksqor1004")
        id = 'lek740815@naver.com'
        pw = '1q2w3e4r5t@'
        self.lineEdit_3.setText(id)
        self.lineEdit_4.setEchoMode(QLineEdit.Password)
        self.lineEdit_4.setText(pw)
        self.first_flag=True
        self.buyMore=False


    def start(self):
        print('11')
        self.startNumber=int(self.lineEdit_5.text())
        self.endNumber = int(self.lineEdit_6.text())
        self.id=self.lineEdit_3.text()
        self.pw=self.lineEdit_4.text()
        self.x = Thread(self,self.auth,self.fname,self.startNumber,self.endNumber,self.id,self.pw)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal2.connect(self.slot2)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal3.connect(self.slot3)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))

    def slot2(self, data2):  # 사용자 정의 시그널1에 connect된 function
        reply = QMessageBox.question(self, 'Message', '추가 구매 하시겠습니까?',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.buyMore=True
            self.x.More()
        else:
            self.buyMore=True
            self.x.Escape()

    def slot3(self, data3):  # 사용자 정의 시그널1에 connect된 function
        print('data3:',data3,type(data3))
        self.lineEdit_5.setText(str(data3+1))


    def find(self):
        print("find")
        self.fname=QFileDialog.getOpenFileName(self," Open file",' ./')[0]
        print(self.fname)
        self.lineEdit_7.setText(self.fname)

    def setSlot(self):
        pass
    def auth(self):

        firebase_id,firebase_password=get_key(self.first_flag)
        self.first_flag = False
        self.id=self.lineEdit.text()
        self.password = self.lineEdit_2.text()


        if firebase_id==self.id and firebase_password==self.password:
            print("로그인성공")
            QMessageBox.information(self, "알림창", "로그인 성공")
            self.auth=True
            self.textEdit.append("로그인 성공")

        else:
            QMessageBox.information(self, "알림창", "로그인 실패")
            print("로그인실패")
            self.textEdit.append("로그인 성공")
            self.auth=False

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())




# firstFlag=True
# fName='list.xlsx'
# startNumber=6
# endNumber=99999
# productList=load_excel(fName,startNumber,endNumber)
# wb=openpyxl.load_workbook(fName)
# ws=wb.active
#
#
# url_login='https://login.coupang.com/login/login.pang?rtnUrl=https%3A%2F%2Fwww.coupang.com%2Fnp%2Fpost%2Flogin%3Fr%3Dhttps%253A%252F%252Fwww.coupang.com%252F'
# browser=chrome_browser(url_login)
#
#
# # id='ljj3347@naver.com'
# # pw='dlwndwo2'
# id='lek740815@naver.com'
# pw='1q2w3e4r5t@'
#
#
# browser.implicitly_wait(3)
# input_id=browser.find_element(By.ID,'login-email-input')
# input_id.send_keys(id)
# time.sleep(0.5)
#
# input_pw=browser.find_element(By.ID,'login-password-input')
# input_pw.send_keys(pw)
# time.sleep(0.5)
#
# btn_login=browser.find_element(By.CSS_SELECTOR,'body > div.member-wrapper.member-wrapper--flex > div.member-main > div.member-login._loginRoot.sms-login-target.style-v2 > form > div.login__content.login__content--trigger > button')
# browser.execute_script("arguments[0].click();", btn_login)  #
#
# while True:
#     print("로그인시도중..")
#     soup=BeautifulSoup(browser.page_source,'lxml')
#     isInputKeyword=len(soup.find_all('input',attrs={'id':'headerSearchKeyword'}))
#     if isInputKeyword>=1:
#         print("로그인완료")
#         break
#     time.sleep(1)
#
#
# for productElem in productList:
#     url_purchase=productElem['url']
#     browser.get(url_purchase)
#     browser.implicitly_wait(3)
#
#
#     try:
#         btnUp=browser.find_element(By.CLASS_NAME,'prod-quantity__plus')
#         actionBtn=productElem['productQuantity']-1
#         print("증가횟수:",actionBtn)
#         for i in range(0,actionBtn):
#             print(f"{i}번째 클릭")
#             btnUp = browser.find_element(By.CLASS_NAME, 'prod-quantity__plus')
#             btnUp.click()
#             time.sleep(3)
#     except:
#         print("갯수 증가 안됨")
#
#     # 구매가능갯수 파악하기
#     soup=BeautifulSoup(browser.page_source,'lxml')
#     scripts=soup.find_all('script')
#     result=""
#     for script in scripts:
#         if str(script).find("wishList")>=0:
#             result=script
#             break
#
#     rawscript=str(result)
#     # print(rawscript)
#     splitPositionFr=rawscript.find("=")
#     splitPositionRr=rawscript.find(";")
#     # print(splitPositionFr,splitPositionRr)
#     rawscriptChanged=rawscript[splitPositionFr+1:splitPositionRr].strip()
#     # print(rawscriptChanged)
#     jsonRawScript=json.loads(rawscriptChanged)
#     # pprint.pprint(jsonRawScript)
#     buyableQuantity=jsonRawScript['buyableQuantity']
#     print("구매가능수량:",buyableQuantity)
#     print("구매할수량:",productElem['productQuantity'])
#
#
#     if buyableQuantity==None:
#         print('구매불가능')
#         ws.cell(row=productElem['rowNum'], column=10).value = "주문불가"
#         wb.save(fName)
#         continue
#     elif buyableQuantity==0 or buyableQuantity<productElem['productQuantity']:
#         print('구매불가능')
#         ws.cell(row=productElem['rowNum'], column=10).value = "주문불가"
#         wb.save(fName)
#         continue
#
#
#     btnPurchase=browser.find_element(By.CLASS_NAME,'prod-buy-btn__txt')
#     btnPurchase.click()
#     print("구매버튼클릭하기")
#     browser.implicitly_wait(5)
#
#     while True:
#         try:
#             btnAddress=browser.find_element(By.CSS_SELECTOR,'#body > div.middle > div:nth-child(4) > h2 > button')
#             btnAddress.click()
#             time.sleep(1)
#             print("구매버튼누르기성공")
#             break
#         except:
#             print("구매버튼누르기실패")
#         time.sleep(1)
#
#
#
#     while True:
#         print("창 뜨는것 대기, 창갯수:{}".format(len(browser.window_handles)))
#         if len(browser.window_handles)>=2:
#
#             break
#         time.sleep(0.5)
#
#     browser.switch_to.window(browser.window_handles[-1])
#     print("마지막창으로 이동")
#
#
#     btnModify=browser.find_element(By.CSS_SELECTOR,'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div > div.address-card.address-card--picked > div.address-card__foot > form.address-card__form.address-card__form--edit._addressBookAddressCardEditForm > button')
#     btnModify.click()
#     browser.implicitly_wait(3)
#
#     btnDelete=browser.find_element(By.CSS_SELECTOR,'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div.content-body__corset > div > form > button')
#     btnDelete.click()
#     browser.implicitly_wait(3)
#
#
#     interval=1
#     prev_height=browser.execute_script('return document.body.scrollHeight')
#     while True:
#         browser.execute_script('window.scrollTo(0,document.body.scrollHeight)')
#         time.sleep(interval)
#         curr_height = browser.execute_script('return document.body.scrollHeight')
#         if curr_height == prev_height:
#             break
#             prev_height=curr_height
#
#
#
#     btnAddAddress=browser.find_element(By.CSS_SELECTOR,'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div > form > div > button')
#     btnAddAddress.click()
#     # browser.execute_script("arguments[0].click();", btnAddAddress)  #
#     time.sleep(1)
#
#
#
#     inputName=browser.find_element(By.ID,'addressbookRecipient')
#     inputName.send_keys(productElem['receiverName'])
#     time.sleep(0.5)
#     print("이름입력")
#
#     inputPhone=browser.find_element(By.ID,'addressbookCellphone')
#     inputPhone.send_keys(productElem['receiverPhone'])
#     time.sleep(0.5)
#     print("전화번호입력")
#
#     searchAddress=browser.find_element(By.CSS_SELECTOR,'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div.content-body__corset > form > div.icon-text-field__frame-box._addressBookAddressErrorStatus > div > div.icon-text-field__button-container > a')
#     searchAddress.click()
#     browser.implicitly_wait(3)
#     print("서칭하기2")
#
#
#     soup=BeautifulSoup(browser.page_source,'lxml')
#     # print(soup.prettify())
#
#     element = browser.find_element(By.CLASS_NAME,"identity__iframe")
#     browser.switch_to.frame(element) #프레임 이동
#
#     inputAddressBasic=browser.find_element(By.NAME,'searchKey')
#     inputAddressBasic.send_keys(productElem['basicAddress'])
#     time.sleep(0.5)
#     print("기본주소정보입력")
#
#
#     btnSearch=browser.find_element(By.CSS_SELECTOR,'body > section > div.zipcode__wrapper > div > div > header > div > form > div.zipcode__search-trigger > button')
#     browser.execute_script("arguments[0].click();", btnSearch)  #
#     browser.implicitly_wait(3)
#     print("서치버튼누르기3")
#     btnRowAddress=browser.find_element(By.CSS_SELECTOR,'body > section > div.zipcode__wrapper > div > div > div > div.zipcode__slide-view._zipcodeResultSlideRoot > div.zipcode__slide-track._zipcodeResultSlide > div.zipcode__slide-item.zipcode__slide-item--address._zipcodeResultSlideItem > div._zipcodeResultListAddress > div:nth-child(1) > span.zipcode__result__item.zipcode__result__item--road._zipcodeResultSendTrigger')
#     browser.execute_script("arguments[0].click();", btnRowAddress)  #
#     browser.implicitly_wait(3)
#     print("첫번째행선택")
#
#
#     browser.switch_to.default_content()
#     soup=BeautifulSoup(browser.page_source,'lxml')
#     # print(soup.prettify())
#     inputDetail=browser.find_element(By.CSS_SELECTOR,'#addressbookAddressDetail')
#     inputDetail.send_keys(productElem['detailAddress'])
#     time.sleep(0.5)
#
#     checkBase=browser.find_element(By.ID,'_addressBookSaveAsDefault')
#     browser.execute_script("arguments[0].click();", checkBase)  #
#     time.sleep(0.5)
#
#     btnSave=browser.find_element(By.CSS_SELECTOR,'body > div > div > div.content-wrapper > div.content-body.content-body--fixed > div.content-body__corset > form > div.addressbook__button-fixer > button')
#     browser.execute_script("arguments[0].click();", btnSave)  #
#     time.sleep(0.5)
#
#     browser.switch_to.window(browser.window_handles[0])
#
#
#     if firstFlag==True:
#         paymentType=browser.find_element(By.CSS_SELECTOR,'#payType9')
#         browser.execute_script("arguments[0].click();", paymentType)  #
#         time.sleep(0.5)
#
#         btnChangeInfo=browser.find_element(By.CSS_SELECTOR,'#body > div.middle > div:nth-child(8) > div > div > button')
#         btnChangeInfo.click()
#         time.sleep(0.5)
#
#         selectOption=browser.find_element(By.CSS_SELECTOR,'#body > div.middle > div:nth-child(8) > div > div > div:nth-child(2) > div.cash-receipt__resiter-type__wrap > span:nth-child(1) > select')
#         selectOption.click()
#         time.sleep(0.5)
#
#         selectSaupja=browser.find_element(By.CSS_SELECTOR,'#body > div.middle > div:nth-child(8) > div > div > div:nth-child(2) > div.cash-receipt__resiter-type__wrap > span:nth-child(1) > select > option:nth-child(2)')
#         selectSaupja.click()
#         time.sleep(0.5)
#
#         inputSaupja=browser.find_element(By.CSS_SELECTOR,'#body > div.middle > div:nth-child(8) > div > div > div:nth-child(2) > div.cash-receipt__resiter-type__wrap > span:nth-child(2) > input')
#
#         inputSaupja.click()
#         time.sleep(0.5)
#         ActionChains(browser).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
#         time.sleep(0.5)
#         ActionChains(browser).send_keys('delete')
#
#         inputSaupja.send_keys(productElem['sellerNo'])
#     firstFlag=False
#
#     btnPay=browser.find_element(By.CSS_SELECTOR,'#paymentBtn')
#     browser.execute_script("arguments[0].click();", btnPay)  #
#     successFlag=False
#     successCount=0
#     while True:
#         soup=BeautifulSoup(browser.page_source,'lxml')
#         if successCount>=30:
#             print("조회 한도 초과로 실패")
#             break
#         # print(soup.prettify())
#         try:
#             payStatus=soup.find('span',attrs={'class':'i18n-wrapper'})
#             payStatusText=payStatus.get_text()
#             print('payStatusText:',payStatusText)
#             if payStatusText.find("완료")>=0:
#                 print("주문완료됨")
#                 successFlag=True
#                 break
#         except:
#             print(f"아직안뜸_{successCount}")
#         successCount+=1
#
#         time.sleep(1)
#
#     if successFlag==True:
#         print("주문완료로 엑셀에 저장")
#         ws.cell(row=productElem['rowNum'],column=10).value="주문성공"
#     else:
#         print("주문실패로 엑셀에 저장")
#         ws.cell(row=productElem['rowNum'], column=10).value = "주문불가"
#     wb.save(fName)
#     determinant=input("또 결제할까요?")
#     if determinant=="Y":
#         print("반복")
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
